"""Excel (openpyxl) report exporters."""

from __future__ import annotations

import io
from typing import Any, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from reports.services import data as report_data

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=14, bold=True)


def _autosize(ws, max_width: int = 40) -> None:
    for column_cells in ws.columns:
        length = 0
        column = column_cells[0].column
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, len(value))
        ws.column_dimensions[get_column_letter(column)].width = min(length + 2, max_width)


def _write_sheet(
    wb: Workbook,
    title: str,
    headers: Sequence[str],
    rows: Iterable[dict[str, Any]],
    keys: Sequence[str],
) -> None:
    ws = wb.create_sheet(title=title[:31])
    ws.append([title])
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.append([])
    ws.append(list(headers))
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append([row.get(key, "") for key in keys])
    _autosize(ws)


def _workbook_bytes(wb: Workbook) -> bytes:
    # Remove default empty sheet if we added named sheets.
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_assets_by_department_excel() -> bytes:
    wb = Workbook()
    _write_sheet(
        wb,
        "By Department",
        [
            "Department",
            "Code",
            "Asset Tag",
            "Asset Name",
            "Category",
            "Status",
            "Employee Code",
            "Employee Email",
            "Assigned At",
        ],
        report_data.assets_by_department_rows(),
        [
            "department",
            "department_code",
            "asset_tag",
            "asset_name",
            "category",
            "status",
            "employee_code",
            "employee_email",
            "assigned_at",
        ],
    )
    _write_sheet(
        wb,
        "Summary",
        ["Department", "Code", "Asset Count"],
        report_data.assets_by_department_summary(),
        ["department", "department_code", "asset_count"],
    )
    return _workbook_bytes(wb)


def export_assets_by_category_excel() -> bytes:
    wb = Workbook()
    _write_sheet(
        wb,
        "By Category",
        [
            "Category",
            "Code",
            "Asset Tag",
            "Asset Name",
            "Status",
            "Serial",
            "Brand",
            "Model",
            "Vendor",
            "Purchase Cost",
        ],
        report_data.assets_by_category_rows(),
        [
            "category",
            "category_code",
            "asset_tag",
            "asset_name",
            "status",
            "serial_number",
            "brand",
            "model",
            "vendor",
            "purchase_cost",
        ],
    )
    _write_sheet(
        wb,
        "Summary",
        ["Category", "Code", "Asset Count"],
        report_data.assets_by_category_summary(),
        ["category", "category_code", "asset_count"],
    )
    return _workbook_bytes(wb)


def export_allocation_history_excel() -> bytes:
    wb = Workbook()
    _write_sheet(
        wb,
        "Allocation History",
        [
            "Asset Tag",
            "Asset Name",
            "Employee Code",
            "Employee Email",
            "Department",
            "Status",
            "Assigned At",
            "Returned At",
            "Assigned By",
            "Notes",
        ],
        report_data.allocation_history_rows(),
        [
            "asset_tag",
            "asset_name",
            "employee_code",
            "employee_email",
            "department",
            "status",
            "assigned_at",
            "returned_at",
            "assigned_by",
            "notes",
        ],
    )
    return _workbook_bytes(wb)


def export_maintenance_history_excel() -> bytes:
    wb = Workbook()
    _write_sheet(
        wb,
        "Maintenance History",
        [
            "Asset Tag",
            "Asset Name",
            "Title",
            "Status",
            "Reported By",
            "Assigned To",
            "Cost",
            "Started At",
            "Completed At",
            "Created At",
            "Resolution",
        ],
        report_data.maintenance_history_rows(),
        [
            "asset_tag",
            "asset_name",
            "title",
            "status",
            "reported_by",
            "assigned_to",
            "cost",
            "started_at",
            "completed_at",
            "created_at",
            "resolution_notes",
        ],
    )
    return _workbook_bytes(wb)
